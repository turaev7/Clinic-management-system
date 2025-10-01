# -*- coding: utf-8 -*-
import os
from datetime import datetime, date, time
from functools import wraps
from sqlalchemy import or_, func


from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)  # ensure folder exists for Excel exports

db = SQLAlchemy(app)

# -------------------- Models --------------------

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(64), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(32), nullable=False)  # 'superadmin' or 'palata'
    ward_access = db.Column(db.String(32), nullable=True)  # e.g., 'palata1', 'palata2', 'palata3'

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class Ward(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(64), nullable=False)  # e.g., 'A-101'
    sort_order = db.Column(db.Integer, nullable=False, default=0)
    block = db.Column(db.String(1), nullable=False, default='A')  # 'A' or 'B'


class Doctor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(128), nullable=False)
    sort_order = db.Column(db.Integer, nullable=False, default=0)


class Patient(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    hist_number = db.Column(db.String(64), nullable=False)   # Ист номер
    last_name = db.Column(db.String(64), nullable=False)     # Фамилия
    first_name = db.Column(db.String(64), nullable=False)    # Имя
    patronymic = db.Column(db.String(64), nullable=False)    # Отечество
    birth_date = db.Column(db.String(10), nullable=False)    # dd.mm.yyyy (text)
    phone = db.Column(db.String(32), nullable=False)
    address = db.Column(db.String(128), nullable=False)
    occupation = db.Column(db.String(64), nullable=False)

    arrival_date = db.Column(db.String(10), nullable=False)  # dd.mm.yyyy
    arrival_time = db.Column(db.String(5), nullable=False)   # HH:MM
    ward_id = db.Column(db.Integer, db.ForeignKey('ward.id'), nullable=False)
    doctor_id = db.Column(db.Integer, db.ForeignKey('doctor.id'), nullable=False)

    caregiver_exists = db.Column(db.Boolean, nullable=False, default=False)
    caregiver_fullname = db.Column(db.String(128), nullable=True)
    caregiver_ward_id = db.Column(db.Integer, db.ForeignKey('ward.id'), nullable=True)
    caregiver_arrival_date = db.Column(db.String(10), nullable=True)
    caregiver_departure_date = db.Column(db.String(10), nullable=True)

    discharge_datetime = db.Column(db.String(16), nullable=True)  # dd.mm.yyyy HH:MM

    ward = db.relationship('Ward', foreign_keys=[ward_id])
    caregiver_ward = db.relationship('Ward', foreign_keys=[caregiver_ward_id])
    doctor = db.relationship('Doctor')

# -------------------- Internationalization --------------------

def get_lang():
    return session.get('lang', 'uz')

I18N = {
    # Uzbek — Latin
    'uz': {
        'app_title': 'Klinika Palata Boshqaruvi',
        'login': 'Kirish',
        'username': 'Login',
        'password': 'Parol',
        'logout': 'Chiqish',
        'registration': 'Ro‘yxatdan o‘tkazish',
        'patients': 'Bemorlar',
        'inpatient': 'Statsionardagi bemorlar',
        'settings': 'Sozlamalar',
        'wards_list': 'Palatalar ro‘yxati',
        'doctors_list': 'Shifokorlar ro‘yxati',
        'import': 'Import',
        'lang': 'Til',
        'save': 'Saqlash',
        'add_new': 'Yangi qo‘shish',
        'delete_selected': 'Belgilanganlarni o‘chirish',
        'name': 'Nomi',
        'sort_order': 'Tartib raqami',
        'block': 'Blok',
        'A_block': 'A blok',
        'B_block': 'B blok',
        'hist_number': 'Istoriya raqami',
        'patient_full_name': 'Bemor F.I.O',
        'last_name': 'Familiyasi',
        'first_name': 'Ismi',
        'patronymic': 'Otasining ismi',
        'birth_date': 'Tug‘ilgan sana',
        'phone': 'Telefon',
        'address': 'Yashash manzili',
        'occupation': 'Kasbi',
        'arrival_date': 'Kelgan sana',
        'arrival_time': 'Kelgan vaqti ',
        'arrival_datetime': 'Kelgan payt',
        'ward': 'Palata',
        'doctor': 'Shifokor',
        'caregiver': 'Qarovchi',
        'yes': 'Ha',
        'no': 'Yo‘q',
        'caregiver_fullname': 'Qarovchi F.I.O',
        'caregiver_ward': 'Palata (qarovchi)',
        'caregiver_arrival_date': 'Kelgan sana (qarovchi)',
        'caregiver_departure_date': 'Ketgan sana (qarovchi)',
        'discharge': 'Ketgan payt ',
        'discharge_datetime': 'Ketgan payt ',
        'search': 'Qidirish',
        'cancel': 'Bekor qilish',
        'export': 'Eksport (Excel)',
        'edit_profile': 'Profilni tahrirlash',
        'caregiver_details': 'Qarovchi ma’lumotlari',
        'submit': 'Yuborish',
        'patients_per_page': 'Har sahifada 500 ta bemor',
        'only_superadmin': 'Bu sahifaga faqat superadmin kira oladi.',
        'patient_fullname': 'Bemor F.I.O',
        'as_of': 'Qaysi payt bo‘yicha',
        'show': 'Ko‘rsatish',
        'instructions_title': 'Ko‘rsatmalar',
        'import_instr_xlsx_only': 'Faqat .xlsx faylini yuklang (Excel 2007+). Eski .xls qo‘llab-quvvatlanmaydi.',
        'import_instr_headers': 'Quyidagi shablondagi sarlavhalardan yoki UZ/RU/EN ekvivalentlaridan foydalaning.',
        'import_instr_required_cols': 'Majburiy ustunlar: Istoriya raqami, Familiyasi, Ismi, Otasining ismi, Tug‘ilgan sana.',
        'download_template': 'Shablonni yuklab olish (.xlsx)',
        'excel_file_label': 'Excel fayli (.xlsx)',
        'import_btn': 'Import',
        'C_block': 'C blok',
        'D_block': 'Diagnostika',
        'R_block': 'Reanimatsiya',
        'as_of': "Vaqtga ko'ra",
        'show': "Ko'rsatish",
        'patients_col': "Bemor(lar)",
        'cleanup_invalid': "Majburiy maydonlari yoʻq bemorlar",
        'cleanup_intro': "Quyidagi bemorlar majburiy maydonlarni toʻliq toʻldirmagan. Pastdagi tugma orqali ularning barchasini oʻchirishingiz mumkin.",
        'missing_fields': "Yetishmayotgan maydonlar",
        'delete_all_listed': "Koʻrsatilganlarning barchasini oʻchirish",
        'back': "Ortga",
        'deleted_n_patients': "{} ta bemor oʻchirildi",
        'no_invalid_records': "Majburiy maydonsiz bemor topilmadi.",


    },
    # Russian
    'ru': {
        'app_title': 'Управление палатами клиники',
        'login': 'Вход',
        'username': 'Логин',
        'password': 'Пароль',
        'logout': 'Выход',
        'registration': 'Регистрация',
        'patients': 'Пациенты',
        'inpatient': 'Пациенты в стационаре',
        'settings': 'Настройки',
        'wards_list': 'Палата лист',
        'doctors_list': 'Врач лист',
        'import': 'Импорт',
        'lang': 'Язык',
        'save': 'Сохранить',
        'add_new': 'Добавить новый',
        'delete_selected': 'Удалить выбранные',
        'name': 'Название',
        'sort_order': 'Порядок',
        'block': 'Блок',
        'A_block': 'A блок',
        'B_block': 'B блок',
        'hist_number': 'Ист номер',
        'patient_full_name': 'Ф.И.О пациента',
        'last_name': 'Фамилия',
        'first_name': 'Имя',
        'patronymic': 'Отчество',
        'birth_date': 'Дата рождения ',
        'phone': 'Тел. номер',
        'address': 'Адрес проживания',
        'occupation': 'Профессия',
        'arrival_date': 'Дата поступления ',
        'arrival_time': 'Время поступления ',
        'arrival_datetime': 'Время поступления',
        'ward': 'Палата',
        'doctor': 'Врач',
        'caregiver': 'Ухажёр',
        'yes': 'Да',
        'no': 'Нет',
        'caregiver_fullname': 'Ухажёр Ф.И.О',
        'caregiver_ward': 'Палата (ухажёр)',
        'caregiver_arrival_date': 'Дата поступления (ухажёр)',
        'caregiver_departure_date': 'Дата выписки (ухажёр)',
        'discharge': 'Дата выписки (дд.мм.гггг ЧЧ:MM)',
        'discharge_datetime': 'Дата выписки',
        'search': 'Поиск',
        'cancel': 'Отменить',
        'export': 'Экспорт (Excel)',
        'edit_profile': 'Редактировать профиль',
        'caregiver_details': 'Данные ухажёр',
        'submit': 'Отправить',
        'patients_per_page': 'По 500 пациентов на странице',
        'only_superadmin': 'Доступ только для супер-админа.',
        'patient_fullname': 'Ф.И.О пациента',
        'as_of': 'На момент',
        'show': 'Показать',
        'instructions_title': 'Инструкции',
        'import_instr_xlsx_only': 'Загрузите файл .xlsx (Excel 2007+). Старый .xls не поддерживается.',
        'import_instr_headers': 'Используйте заголовки из шаблона ниже или их UZ/RU/EN эквиваленты.',
        'import_instr_required_cols': 'Обязательные столбцы: Ист номер, Фамилия, Имя, Отчество, Дата рождения.',
        'download_template': 'Скачать шаблон (.xlsx)',
        'excel_file_label': 'Файл Excel (.xlsx)',
        'import_btn': 'Импорт',
        'C_block': 'C блок',
        'D_block': 'Диагностика',
        'R_block': 'Реанимация',
        'as_of': "На дату/время",
        'show': "Показать",
        'patients_col': "Пациент(ы)",
        'cleanup_invalid': "Пациенты с пустыми обязательными полями",
        'cleanup_intro': "У следующих пациентов не заполнены обязательные поля. Вы можете удалить их все кнопкой ниже.",
        'missing_fields': "Отсутствующие поля",
        'delete_all_listed': "Удалить всех в списке",
        'back': "Назад",
        'deleted_n_patients': "Удалено {} пациентов",
        'no_invalid_records': "Пациенты без обязательных полей не найдены.",


    },
    # English
    'en': {
        'app_title': 'Clinic Ward Management',
        'login': 'Login',
        'username': 'Username',
        'password': 'Password',
        'logout': 'Logout',
        'registration': 'Registration',
        'patients': 'Patients',
        'inpatient': 'Inpatients (By Ward)',
        'settings': 'Settings',
        'wards_list': 'Ward List',
        'doctors_list': 'Doctor List',
        'import': 'Import',
        'lang': 'Language',
        'save': 'Save',
        'add_new': 'Add New',
        'delete_selected': 'Delete Selected',
        'name': 'Name',
        'sort_order': 'Sort Order',
        'block': 'Block',
        'A_block': 'Block A',
        'B_block': 'Block B',
        'hist_number': 'History No.',
        'patient_full_name': 'Patient Full Name',
        'last_name': 'Last Name',
        'first_name': 'First Name',
        'patronymic': 'Patronymic',
        'birth_date': 'Date of Birth (dd.mm.yyyy)',
        'phone': 'Phone',
        'address': 'Address',
        'occupation': 'Occupation',
        'arrival_date': 'Arrival Date (dd.mm.yyyy)',
        'arrival_time': 'Arrival Time (HH:MM)',
        'arrival_datetime': 'Arrival DateTime',
        'ward': 'Ward',
        'doctor': 'Doctor',
        'caregiver': 'Caregiver',
        'yes': 'Yes',
        'no': 'No',
        'caregiver_fullname': 'Caregiver Full Name',
        'caregiver_ward': 'Ward (Caregiver)',
        'caregiver_arrival_date': 'Arrival Date (Caregiver)',
        'caregiver_departure_date': 'Departure Date (Caregiver)',
        'discharge': 'Discharge (dd.mm.yyyy HH:MM)',
        'discharge_datetime': 'Discharge time',
        'search': 'Search',
        'cancel': 'Cancel',
        'export': 'Export (Excel)',
        'edit_profile': 'Edit Profile',
        'caregiver_details': 'Caregiver Details',
        'submit': 'Submit',
        'patients_per_page': '500 patients per page',
        'only_superadmin': 'Only superadmin can access this page.',
        'patient_fullname': 'Patient Full Name',
        'as_of': 'As of',
        'show': 'Show',
        'instructions_title': 'Instructions',
        'import_instr_xlsx_only': 'Upload an .xlsx file (Excel 2007+). Old .xls is not supported.',
        'import_instr_headers': 'Use the header names from the template below, or equivalent Uzbek/Russian/English variants.',
        'import_instr_required_cols': 'Minimum required columns: History No., Last Name, First Name, Patronymic, Date of Birth.',
        'download_template': 'Download template (.xlsx)',
        'excel_file_label': 'Excel file (.xlsx)',
        'import_btn': 'Import',
        'C_block': 'Block C',
        'D_block': 'Diagnostics',
        'R_block': 'Reanimation',
        'as_of': "As of",
        'show': "Show",
        'patients_col': "Patient(s)",
        'cleanup_invalid': "Patients with missing required fields",
        'cleanup_intro': "The patients below are missing required fields. You can delete them all using the button below.",
        'missing_fields': "Missing fields",
        'delete_all_listed': "Delete all listed",
        'back': "Back",
        'deleted_n_patients': "Deleted {} patients",
        'no_invalid_records': "No patients with missing required fields.",
    }
}

def t(key):
    lang = get_lang()
    return I18N.get(lang, I18N['uz']).get(key, key)

# -------------------- Helpers --------------------

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapper

def superadmin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        user = User.query.get(session['user_id'])
        if not user or user.role != 'superadmin':
            flash(t('only_superadmin'), 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return wrapper

def init_db():
    db.create_all()
    # Seed users if not exist
    if User.query.count() == 0:
        users = [
            ('superadmin', '5358287', 'superadmin', None),
            ('palata1', '7825065', 'palata', 'palata1'),
            ('palata2', '6037809', 'palata', 'palata2'),
            ('palata3', '7095823', 'palata', 'palata3'),
        ]
        for u, p, r, wa in users:
            db.session.add(User(username=u, password_hash=generate_password_hash(p), role=r, ward_access=wa))
        db.session.commit()
    # Seed example wards/doctors
    if Ward.query.count() == 0:
        for i in range(1, 6):
            db.session.add(Ward(name=f"A-{100+i}", sort_order=i, block='A'))
        for i in range(1, 6):
            db.session.add(Ward(name=f"B-{200+i}", sort_order=100+i, block='B'))
        db.session.commit()
    if Doctor.query.count() == 0:
        for i in range(1, 6):
            db.session.add(Doctor(full_name=f"Dr. Example {i}", sort_order=i))
        db.session.commit()

@app.before_request
def ensure_init():
    # Create DB on first run
    if not os.path.exists(os.path.join(os.path.dirname(__file__), 'database.db')):
        with app.app_context():
            init_db()

# -------- Snapshot helpers (for historical occupancy on /inpatient) --------

def _parse_ddmmyyyy(s):
    try:
        d, m, y = s.strip().split('.')
        return int(y), int(m), int(d)
    except Exception:
        return None

def _parse_hhmm(s):
    try:
        h, m = s.strip().split(':')
        return int(h), int(m)
    except Exception:
        return 0, 0

def _to_dt(date_str, time_str=None):
    """date_str = 'dd.mm.yyyy', time_str = 'HH:MM' or None"""
    if not date_str:
        return None
    ymd = _parse_ddmmyyyy(date_str)
    if not ymd:
        return None
    Y, M, D = ymd
    H, Mi = _parse_hhmm(time_str or '00:00')
    try:
        return datetime(Y, M, D, H, Mi)
    except Exception:
        return None

def _parse_discharge(discharge_str):
    """Accepts 'dd.mm.yyyy HH:MM' OR 'dd.mm.yyyy'."""
    if not discharge_str:
        return None
    parts = discharge_str.strip().split()
    if len(parts) == 1:
        return _to_dt(parts[0], '00:00')
    if len(parts) >= 2:
        return _to_dt(parts[0], parts[1])
    return None

def _is_active_at(patient, ref_dt):
    """True if patient was in ward at ref_dt (arrival <= ref_dt < discharge OR no discharge)."""
    arr = _to_dt(patient.arrival_date, patient.arrival_time)
    if not arr or arr > ref_dt:
        return False
    dis = _parse_discharge(patient.discharge_datetime)
    if dis and dis <= ref_dt:
        return False
    return True

# -------------------- Routes --------------------

@app.route('/set_lang/<lang>')
def set_lang(lang):
    if lang in I18N:
        session['lang'] = lang
    next_url = request.args.get('next')  # stay on the same page & keep query (?at=...)
    return redirect(next_url or request.referrer or url_for('index'))

@app.route('/')
@login_required
def index():
    return redirect(url_for('register'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            session['user_id'] = user.id
            return redirect(url_for('index'))
        flash('Invalid credentials', 'danger')
    return render_template('login.html', t=t)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

def _invalid_required_conditions():
    """SQLAlchemy conditions to find patients missing any required field."""
    def is_blank(col):
        return or_(col.is_(None), func.length(func.trim(col)) == 0)

    return [
        is_blank(Patient.hist_number),
        is_blank(Patient.last_name),
        is_blank(Patient.first_name),
        is_blank(Patient.patronymic),
        is_blank(Patient.birth_date),
        is_blank(Patient.phone),
        is_blank(Patient.address),
        is_blank(Patient.occupation),
        is_blank(Patient.arrival_date),
        is_blank(Patient.arrival_time),
        Patient.ward_id.is_(None),
        Patient.doctor_id.is_(None),
    ]

# -------------------- Registration --------------------

@app.route('/register', methods=['GET', 'POST'])
@login_required
def register():
    wards = Ward.query.order_by(Ward.sort_order).all()
    doctors = Doctor.query.order_by(Doctor.sort_order).all()
    if request.method == 'POST':
        data = request.form
        required_fields = [
            'hist_number', 'last_name', 'first_name', 'patronymic', 'birth_date',
            'phone', 'address', 'occupation', 'arrival_date', 'arrival_time', 'ward_id', 'doctor_id'
        ]
        for f in required_fields:
            if not data.get(f):
                flash(f"Missing field: {f}", 'danger')
                return redirect(url_for('register'))
        caregiver_exists = (data.get('caregiver_exists') == 'yes')
        p = Patient(
            hist_number=data.get('hist_number').strip(),
            last_name=data.get('last_name').strip(),
            first_name=data.get('first_name').strip(),
            patronymic=data.get('patronymic').strip(),
            birth_date=data.get('birth_date').strip(),
            phone=data.get('phone').strip(),
            address=data.get('address').strip(),
            occupation=data.get('occupation').strip(),
            arrival_date=data.get('arrival_date').strip(),
            arrival_time=data.get('arrival_time').strip(),
            ward_id=int(data.get('ward_id')),
            doctor_id=int(data.get('doctor_id')),
            caregiver_exists=caregiver_exists,
            caregiver_fullname=(data.get('caregiver_fullname') or '').strip() if caregiver_exists else None,
            caregiver_ward_id=int(data.get('caregiver_ward_id')) if caregiver_exists and data.get('caregiver_ward_id') else None,
            caregiver_arrival_date=(data.get('caregiver_arrival_date') or '').strip() if caregiver_exists else None,
            caregiver_departure_date=(data.get('caregiver_departure_date') or '').strip() if caregiver_exists else None,
            discharge_datetime=(data.get('discharge_datetime') or '').strip() or None
        )
        db.session.add(p)
        db.session.commit()
        flash('Saved', 'success')
        return redirect(url_for('patients'))
    return render_template('register.html', t=t, wards=wards, doctors=doctors)

# -------------------- Patients List --------------------

@app.route('/patients')
@login_required
def patients():
    q_hist = request.args.get('q_hist', '').strip()
    q_last = request.args.get('q_last', '').strip()
    q_first = request.args.get('q_first', '').strip()
    q_pat = request.args.get('q_pat', '').strip()

    query = Patient.query

    if q_hist:
        query = query.filter(Patient.hist_number.ilike(f"{q_hist}%"))
    if q_last:
        query = query.filter(Patient.last_name.ilike(f"{q_last}%"))
    if q_first:
        query = query.filter(Patient.first_name.ilike(f"{q_first}%"))
    if q_pat:
        query = query.filter(Patient.patronymic.ilike(f"{q_pat}%"))

    page = int(request.args.get('page', 1))
    per_page = 500
    pagination = query.order_by(Patient.id.desc()).paginate(page=page, per_page=per_page, error_out=False)
    rows = pagination.items

    wards = {w.id: w for w in Ward.query.all()}
    doctors = {d.id: d for d in Doctor.query.all()}

    return render_template('patients.html', t=t, rows=rows, wards=wards, doctors=doctors, pagination=pagination,
                           q_hist=q_hist, q_last=q_last, q_first=q_first, q_pat=q_pat)

@app.route('/patients/export')
@login_required
def patients_export():
    from sqlalchemy import func

    # current filters
    q_hist = (request.args.get('q_hist') or '').strip()
    q_last = (request.args.get('q_last') or '').strip()
    q_first = (request.args.get('q_first') or '').strip()
    q_pat = (request.args.get('q_pat') or '').strip()

    query = Patient.query
    if q_hist:
        query = query.filter(func.lower(Patient.hist_number).like(f"{q_hist.lower()}%"))
    if q_last:
        query = query.filter(func.lower(Patient.last_name).like(f"{q_last.lower()}%"))
    if q_first:
        query = query.filter(func.lower(Patient.first_name).like(f"{q_first.lower()}%"))
    if q_pat:
        query = query.filter(func.lower(Patient.patronymic).like(f"{q_pat.lower()}%"))

    wards = {w.id: w for w in Ward.query.all()}
    doctors = {d.id: d for d in Doctor.query.all()}

    wb = Workbook()
    ws = wb.active
    ws.title = t('patients')  # sheet title localized

    # Localized headers
    headers = [
        t('hist_number'),
        t('patient_full_name'),
        t('birth_date'),
        t('phone'),
        t('address'),
        t('occupation'),
        t('arrival_datetime'),
        t('discharge_datetime'),
        t('ward'),
        t('doctor'),
        t('caregiver'),
    ]
    ws.append(headers)

    # Rows
    for p in query.all():
        ward = wards.get(p.ward_id)
        doc = doctors.get(p.doctor_id)
        fio = f"{p.last_name} {p.first_name} {p.patronymic}".strip()
        arrive = f"{(p.arrival_date or '').strip()} {(p.arrival_time or '').strip()}".strip()
        caregiver_txt = t('yes') if p.caregiver_exists else t('no')
        ws.append([
            p.hist_number,
            fio,
            p.birth_date,
            p.phone,
            p.address,
            p.occupation,
            arrive,
            p.discharge_datetime or '',
            ward.name if ward else '',
            doc.full_name if doc else '',
            caregiver_txt
        ])

    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    out_path = os.path.join(
        app.config['UPLOAD_FOLDER'],
        f"patients_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    wb.save(out_path)
    return send_file(out_path, as_attachment=True, download_name='patients_export.xlsx')

@app.route('/patients/<int:pid>/edit', methods=['GET', 'POST'])
@login_required
def edit_patient(pid):
    p = Patient.query.get_or_404(pid)
    wards = Ward.query.order_by(Ward.sort_order).all()
    doctors = Doctor.query.order_by(Doctor.sort_order).all()

    if request.method == 'POST':
        data = request.form
        p.hist_number = data.get('hist_number').strip()
        p.last_name = data.get('last_name').strip()
        p.first_name = data.get('first_name').strip()
        p.patronymic = data.get('patronymic').strip()
        p.birth_date = data.get('birth_date').strip()
        p.phone = data.get('phone').strip()
        p.address = data.get('address').strip()
        p.occupation = data.get('occupation').strip()
        p.arrival_date = data.get('arrival_date').strip()
        p.arrival_time = data.get('arrival_time').strip()
        p.ward_id = int(data.get('ward_id'))
        p.doctor_id = int(data.get('doctor_id'))
        p.caregiver_exists = (data.get('caregiver_exists') == 'yes')
        p.caregiver_fullname = (data.get('caregiver_fullname') or '').strip() if p.caregiver_exists else None
        p.caregiver_ward_id = int(data.get('caregiver_ward_id')) if p.caregiver_exists and data.get('caregiver_ward_id') else None
        p.caregiver_arrival_date = (data.get('caregiver_arrival_date') or '').strip() if p.caregiver_exists else None
        p.caregiver_departure_date = (data.get('caregiver_departure_date') or '').strip() if p.caregiver_exists else None
        p.discharge_datetime = (data.get('discharge_datetime') or '').strip() or None

        db.session.commit()
        flash('Saved', 'success')
        return redirect(url_for('patients'))
    return render_template('edit_patient.html', t=t, p=p, wards=wards, doctors=doctors)

# -------------------- Inpatient by Ward (with historical snapshot) --------------------

def _parse_dt(date_str, time_str=None):
    ds = (date_str or '').strip()
    ts = (time_str or '').strip()
    if not ds:
        return None
    try:
        d = datetime.strptime(ds, "%d.%m.%Y")
    except Exception:
        return None
    if ts:
        try:
            t = datetime.strptime(ts, "%H:%M").time()
        except Exception:
            t = time.min
    else:
        t = time.min
    return datetime.combine(d.date(), t)

@app.route('/inpatient')
@login_required
def inpatient():
    at_str = (request.args.get('at') or '').strip()  # "dd.mm.yyyy HH:MM"
    at_dt = None
    if at_str:
        try:
            at_dt = datetime.strptime(at_str, "%d.%m.%Y %H:%M")
        except Exception:
            try:
                at_dt = datetime.strptime(at_str, "%d.%m.%Y")
            except Exception:
                at_dt = None
    if not at_dt:
        at_dt = datetime.now()
        at_str = at_dt.strftime("%d.%m.%Y %H:%M")

    wards = Ward.query.order_by(Ward.block.asc(), Ward.sort_order.asc()).all()

    # Build ward -> occupants at at_dt
    ward_patients = {w.id: [] for w in wards}

    for p in Patient.query.all():
        arrive_dt = _parse_dt(p.arrival_date, p.arrival_time)
        if not arrive_dt or arrive_dt > at_dt:
            continue

        # discharge parsing
        disc_dt = None
        if p.discharge_datetime:
            s = p.discharge_datetime.strip()
            for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y"):
                try:
                    disc_dt = datetime.strptime(s, fmt)
                    break
                except Exception:
                    pass

        if not disc_dt or disc_dt >= at_dt:
            ward_patients.setdefault(p.ward_id, []).append({
                "name": f"{p.last_name} {p.first_name} {p.patronymic}".strip(),
                "hist": p.hist_number,
                "type": "patient"
            })

        # Caregiver occupies a bed too
        if p.caregiver_exists:
            cg_ward_id = p.caregiver_ward_id or p.ward_id
            cg_arrive = _parse_dt(p.caregiver_arrival_date, None)
            cg_depart = _parse_dt(p.caregiver_departure_date, None)
            in_time = True
            if cg_arrive and cg_arrive > at_dt:
                in_time = False
            if cg_depart and cg_depart < at_dt:
                in_time = False
            if in_time and cg_ward_id in ward_patients:
                cg_name = (p.caregiver_fullname or "").strip()  # may be empty
                ward_patients[cg_ward_id].append({
                    "name": cg_name,
                    "hist": "",
                    "type": "caregiver"
                })

    # 5 blocks
    blocks_order = ['A', 'B', 'C', 'D', 'R']
    blocks = {b: [] for b in blocks_order}
    for w in wards:
        if w.block in blocks:
            blocks[w.block].append(w)

    return render_template('inpatient.html', t=t, wards=wards,
                           ward_patients=ward_patients, blocks=blocks, at_str=at_str)

@app.route('/inpatient/export')
@login_required
def inpatient_export():
    at_str = (request.args.get('at') or '').strip()
    try:
        at_dt = datetime.strptime(at_str, "%d.%m.%Y %H:%M") if at_str else datetime.now()
    except Exception:
        try:
            at_dt = datetime.strptime(at_str, "%d.%m.%Y")
        except Exception:
            at_dt = datetime.now()
    at_str = at_dt.strftime("%d.%m.%Y %H:%M")

    wards = Ward.query.order_by(Ward.block.asc(), Ward.sort_order.asc()).all()

    ward_patients = {w.id: [] for w in wards}
    for p in Patient.query.all():
        arrive_dt = _parse_dt(p.arrival_date, p.arrival_time)
        if not arrive_dt or arrive_dt > at_dt:
            continue

        disc_dt = None
        if p.discharge_datetime:
            s = p.discharge_datetime.strip()
            for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y"):
                try:
                    disc_dt = datetime.strptime(s, fmt)
                    break
                except Exception:
                    pass

        if not disc_dt or disc_dt >= at_dt:
            ward_patients.setdefault(p.ward_id, []).append({
                "name": f"{p.last_name} {p.first_name} {p.patronymic}".strip(),
                "hist": p.hist_number,
                "type": "patient"
            })

        if p.caregiver_exists:
            cg_ward_id = p.caregiver_ward_id or p.ward_id
            cg_arrive = _parse_dt(p.caregiver_arrival_date, None)
            cg_depart = _parse_dt(p.caregiver_departure_date, None)
            in_time = True
            if cg_arrive and cg_arrive > at_dt: in_time = False
            if cg_depart and cg_depart < at_dt: in_time = False
            if in_time and cg_ward_id in ward_patients:
                cg_name = (p.caregiver_fullname or "").strip()
                ward_patients[cg_ward_id].append({
                    "name": cg_name,
                    "hist": "",
                    "type": "caregiver"
                })

    blocks_order = ['A', 'B', 'C', 'D', 'R']
    block_titles = {
        'A': t('A_block'),
        'B': t('B_block'),
        'C': t('C_block'),
        'D': t('D_block'),
        'R': t('R_block'),
    }
    blocks = {b: [] for b in blocks_order}
    for w in wards:
        if w.block in blocks:
            blocks[w.block].append(w)

    wb = Workbook()
    ws = wb.active
    ws.title = t('inpatient')

    ws.append([f"{t('inpatient')} — {at_str}"])
    ws.append([])

    for b in blocks_order:
        ws.append([block_titles[b]])
        ws.append([t('ward'), t('patients_col')])
        for w in blocks[b]:
            occupants = ward_patients.get(w.id, [])
            if occupants:
                lines = []
                for o in occupants:
                    # caregiver label appears only if name exists; otherwise we show just the role later
                    base = o['name'] if o['name'] else t('caregiver') if o['type'] == 'caregiver' else ''
                    suffix = f" ({t('caregiver')})" if (o['type'] == 'caregiver' and o['name']) else ""
                    hist = f" ({o['hist']})" if o['hist'] else ""
                    label = (base or '').strip()
                    if not label and o['type'] != 'caregiver':
                        label = ''
                    lines.append(f"{label}{suffix}{hist}".strip() or "—")
                joined = "\n".join(lines)
            else:
                joined = "—"
            ws.append([w.name, joined])
        ws.append([])

    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    out_path = os.path.join(app.config['UPLOAD_FOLDER'], f"inpatient_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(out_path)
    return send_file(out_path, as_attachment=True, download_name='inpatient_export.xlsx')

# -------------------- Settings (Superadmin only) --------------------

@app.route('/settings')
@superadmin_required
def settings_home():
    return render_template('settings.html', t=t)

@app.route('/settings/wards', methods=['GET', 'POST'])
@superadmin_required
def settings_wards():
    if request.method == 'POST':
        if 'add' in request.form:
            name = request.form.get('name', '').strip()
            sort_order = int(request.form.get('sort_order') or 0)
            block = request.form.get('block', 'A')
            if name:
                w = Ward(name=name, sort_order=sort_order, block=block)
                db.session.add(w)
                db.session.commit()
        elif 'delete_selected' in request.form:
            ids = request.form.getlist('delete_id')
            for id_ in ids:
                w = Ward.query.get(int(id_))
                if w:
                    db.session.delete(w)
            db.session.commit()
        elif 'inline_save' in request.form:
            wid = int(request.form.get('wid'))
            w = Ward.query.get(wid)
            if w:
                w.name = request.form.get('name').strip()
                w.sort_order = int(request.form.get('sort_order') or 0)
                w.block = request.form.get('block', 'A')
                db.session.commit()

    wards = Ward.query.order_by(Ward.sort_order).all()
    return render_template('settings_wards.html', t=t, wards=wards)

@app.route('/settings/clear_patients', methods=['POST'])
@superadmin_required
def settings_clear_patients():
    deleted = db.session.query(Patient).delete()
    db.session.commit()
    flash(f'Deleted {deleted} patients', 'success')
    return redirect(url_for('settings_home'))

@app.route('/settings/doctors', methods=['GET', 'POST'])
@superadmin_required
def settings_doctors():
    if request.method == 'POST':
        if 'add' in request.form:
            full_name = request.form.get('full_name', '').strip()
            sort_order = int(request.form.get('sort_order') or 0)
            if full_name:
                d = Doctor(full_name=full_name, sort_order=sort_order)
                db.session.add(d)
                db.session.commit()
        elif 'delete_selected' in request.form:
            ids = request.form.getlist('delete_id')
            for id_ in ids:
                d = Doctor.query.get(int(id_))
                if d:
                    db.session.delete(d)
            db.session.commit()
        elif 'inline_save' in request.form:
            did = int(request.form.get('did'))
            d = Doctor.query.get(did)
            if d:
                d.full_name = request.form.get('full_name').strip()
                d.sort_order = int(request.form.get('sort_order') or 0)
                db.session.commit()

    doctors = Doctor.query.order_by(Doctor.sort_order).all()
    return render_template('settings_doctors.html', t=t, doctors=doctors)

@app.route('/settings/cleanup-invalid', methods=['GET', 'POST'])
@superadmin_required
def settings_cleanup_invalid():
    # Build query of invalid patients
    conds = _invalid_required_conditions()
    base_q = Patient.query.filter(or_(*conds))

    if request.method == 'POST':
        # Delete ONLY selected
        if 'delete_selected' in request.form:
            raw_ids = request.form.getlist('sel')
            ids = []
            for s in raw_ids:
                try:
                    ids.append(int(s))
                except Exception:
                    pass

            if not ids:
                # Fallback message if i18n key not present
                msg = t('no_selection') if t('no_selection') != 'no_selection' else 'No rows selected.'
                flash(msg, 'warning')
                return redirect(url_for('settings_cleanup_invalid'))

            count = Patient.query.filter(
                Patient.id.in_(ids),
                or_(*conds)  # keep it safe: only delete if still invalid
            ).delete(synchronize_session=False)
            db.session.commit()

            # Localized flash
            try:
                flash(t('deleted_n_patients').format(count), 'success')
            except Exception:
                flash(f'Deleted {count} patients', 'success')

            return redirect(url_for('settings_cleanup_invalid'))

        # Delete ALL listed (same as before)
        if 'delete_all' in request.form:
            count = base_q.delete(synchronize_session=False)
            db.session.commit()
            try:
                flash(t('deleted_n_patients').format(count), 'success')
            except Exception:
                flash(f'Deleted {count} patients', 'success')
            return redirect(url_for('settings_cleanup_invalid'))

    # ----- GET: preview list with "missing fields" -----
    invalids = base_q.all()

    def missing_list(p):
        m = []
        if not p.hist_number or not p.hist_number.strip(): m.append(t('hist_number'))
        if not p.last_name or not p.last_name.strip(): m.append(t('last_name'))
        if not p.first_name or not p.first_name.strip(): m.append(t('first_name'))
        if not p.patronymic or not p.patronymic.strip(): m.append(t('patronymic'))
        if not p.birth_date or not p.birth_date.strip(): m.append(t('birth_date'))
        if not p.phone or not p.phone.strip(): m.append(t('phone'))
        if not p.address or not p.address.strip(): m.append(t('address'))
        if not p.occupation or not p.occupation.strip(): m.append(t('occupation'))
        if not p.arrival_date or not p.arrival_date.strip(): m.append(t('arrival_date'))
        if not p.arrival_time or not p.arrival_time.strip(): m.append(t('arrival_time'))
        if not p.ward_id: m.append(t('ward'))
        if not p.doctor_id: m.append(t('doctor'))
        return m

    rows = [{'p': p, 'missing': missing_list(p)} for p in invalids]
    return render_template('settings_cleanup_invalid.html', t=t, rows=rows)

@app.route('/settings/import', methods=['GET', 'POST'])
@superadmin_required
def settings_import():
    import re
    from datetime import datetime
    from openpyxl import load_workbook

    def flash_back(msg, cat='danger'):
        flash(msg, cat)
        return redirect(url_for('settings_import'))

    # --- helpers -------------------------------------------------------------
    def norm(s):
        s = (s or '').strip().lower()
        return ''.join(ch for ch in s if ch.isalnum())

    # dd.mm.yyyy, dd.mm.yy, dd/mm, yyyy-mm-dd, Excel date -> 'dd.mm.yyyy'
    def parse_date(val):
        if val is None or str(val).strip() == '':
            return ''
        if isinstance(val, datetime):
            return val.strftime('%d.%m.%Y')
        s = str(val).strip()
        # dd.mm or dd/mm -> assume current year
        m = re.match(r'^(\d{1,2})[./-](\d{1,2})$', s)
        if m:
            d, mo = int(m.group(1)), int(m.group(2))
            y = datetime.now().year
            try:
                return datetime(y, mo, d).strftime('%d.%m.%Y')
            except ValueError:
                return s
        # dd.mm.yy / dd.mm.yyyy (or with / or -)
        m = re.match(r'^(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})', s)
        if m:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if y < 100: y += 2000
            try:
                return datetime(y, mo, d).strftime('%d.%m.%Y')
            except ValueError:
                return s
        # yyyy-mm-dd
        m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
        if m:
            try:
                return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).strftime('%d.%m.%Y')
            except ValueError:
                return s
        return s

    # -> 'HH:MM'
    def parse_time(val):
        if val is None or str(val).strip() == '':
            return ''
        if isinstance(val, datetime):
            return val.strftime('%H:%M')
        s = str(val).strip()
        m = re.search(r'(\d{1,2}):(\d{2})', s)
        if m:
            hh = int(m.group(1)); mm = int(m.group(2))
            return f'{hh:02d}:{mm:02d}'
        m = re.match(r'^(\d{2})(\d{2})$', s)
        if m:
            return f'{m.group(1)}:{m.group(2)}'
        return s

    # split "Бемор Ф.И.О" -> (last, first, patronymic)
    def parse_fio(val):
        if not val:
            return '', '', ''
        s = str(val).replace(',', ' ').strip()
        parts = [p for p in s.split() if p]
        if len(parts) >= 3:
            return parts[0], parts[1], ' '.join(parts[2:])
        if len(parts) == 2:
            return parts[0], parts[1], ''
        if len(parts) == 1:
            return parts[0], '', ''
        return '', '', ''

    def cell(row, i):
        if i is None: return ''
        return '' if i >= len(row) or row[i] is None else str(row[i]).strip()

    # --- main ---------------------------------------------------------------
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or f.filename.strip() == '':
            return flash_back('No file selected')
        if not f.filename.lower().endswith('.xlsx'):
            return flash_back('Please upload an .xlsx file (Excel 2007+).')

        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        path = os.path.join(app.config['UPLOAD_FOLDER'], f"import_{ts}.xlsx")
        f.save(path)

        try:
            wb = load_workbook(path, data_only=True)
        except Exception as e:
            return flash_back(f'Could not read Excel: {e}')

        # Prefer "PalataQabul" if present, else active sheet
        ws = wb['PalataQabul'] if 'PalataQabul' in wb.sheetnames else wb.active

        # find header row (first row with >= 6 non-empty cells)
        header_row_idx = 1
        headers = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            vals = [v for v in row if v not in (None, '')]
            if len(vals) >= 6:
                header_row_idx = i
                headers = [str(x).strip() if x is not None else '' for x in row]
                break
        if headers is None:
            header_row_idx = 1
            headers = [str(x).strip() if x is not None else '' for x in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

        headers_norm = [norm(h) for h in headers]

        # Map exact headers from your file (Cyrillic Uzbek/Russian) and fallbacks
        colmap = {
            'hist':      ['тартибраками', 'истномер', 'hist_number', 'historyno', 'номер'],
            'fio':       ['беморфио', 'фио', 'фипациента', 'full_name', 'fio', 'фиопасиента'],
            'dob':       ['тугилгансана', 'датарождения', 'birthdate', 'dateofbirth'],
            'address':   ['доимийяшашжойиёкикариндошякинларинингманзилителефон', 'адрес', 'address'],
            'phone':     ['телефонраками', 'телномер', 'phone', 'телефон'],
            'occ':       ['ишжойи', 'касби', 'occupation', 'profession'],
            'arr_date':  ['келгансана', 'датапоступления', 'arrivaldate'],
            'arr_time':  ['келганвакти', 'времявступления', 'arrivaltime'],
            'disc_date': ['чикарилгансана', 'датавыписки', 'dischargedate'],
            'disc_time': ['чикарилганвакт', 'времявыписки', 'dischargetime'],
            'ward':      ['палата', 'ward', 'palata'],
            'doctor':    ['шифокор', 'врач', 'doctor'],
            'caregiver': ['каровчи', 'сиделка', 'caregiver'],
            # extra (ignored if not in your DB): diagnosis/referrer/reject reason
            'diagnosis': ['кабулхонаташхиси'],
            'referrer':  ['кайсимуассасайуллаганьокикимолибкелган'],
            'reject':    ['раdetишнингсабабиташхис', 'радэтишнингсабабиташхис'],
        }

        def find_col(candidates):
            cands = [norm(c) for c in candidates]
            for i, h in enumerate(headers_norm):
                if h in cands:
                    return i
            for i, h in enumerate(headers_norm):
                if any(h.startswith(c) for c in cands):
                    return i
            return None

        idx = {k: find_col(v) for k, v in colmap.items()}

        # Required: history no (we map "Тартиб Раками") + DOB + FIO
        if idx.get('hist') is None or idx.get('dob') is None or idx.get('fio') is None:
            missing = []
            if idx.get('hist') is None: missing.append('Тартиб Раками')
            if idx.get('dob') is None:  missing.append('Тугилган Сана')
            if idx.get('fio') is None:  missing.append('Бемор Ф.И.О')
            return flash_back('Missing required columns: ' + ', '.join(missing))

        # lookups for ward / doctor names already in DB
        ward_by_name   = {w.name.strip(): w for w in Ward.query.all()}
        doctor_by_name = {d.full_name.strip(): d for d in Doctor.query.all()}

        imported = 0
        for row in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
            row = list(row) if row else []
            hist = cell(row, idx['hist'])
            fio_val = cell(row, idx['fio'])
            dob_raw = row[idx['dob']] if idx['dob'] is not None and idx['dob'] < len(row) else ''
            dob = parse_date(dob_raw)

            if not (hist and fio_val and dob):
                continue

            last, first, pat = parse_fio(fio_val)
            if not (last and first):
                continue  # need at least lastname + firstname

            addr  = cell(row, idx['address'])
            phone = cell(row, idx['phone'])
            occ   = cell(row, idx['occ'])

            ad = parse_date(row[idx['arr_date']]) if idx.get('arr_date') is not None and idx['arr_date'] < len(row) else ''
            at = parse_time(row[idx['arr_time']]) if idx.get('arr_time') is not None and idx['arr_time'] < len(row) else ''

            dd = parse_date(row[idx['disc_date']]) if idx.get('disc_date') is not None and idx['disc_date'] < len(row) else ''
            dt = parse_time(row[idx['disc_time']]) if idx.get('disc_time') is not None and idx['disc_time'] < len(row) else ''
            discharge_dt = f"{dd} {dt}".strip() if (dd or dt) else ''

            ward_name = cell(row, idx['ward'])
            doc_name  = cell(row, idx['doctor'])
            care_val  = cell(row, idx['caregiver'])
            cg_norm = norm(care_val)
            caregiver_exists = cg_norm in ('да','ha','yes','1','true','bor')

            # resolve ward / doctor
            ward_id = ward_by_name.get(ward_name).id if ward_name in ward_by_name else (Ward.query.first().id if Ward.query.first() else None)
            doctor_id = doctor_by_name.get(doc_name).id if doc_name in doctor_by_name else (Doctor.query.first().id if Doctor.query.first() else None)

            p = Patient(
                hist_number=hist,
                last_name=last,
                first_name=first,
                patronymic=pat,
                birth_date=dob,
                phone=phone,
                address=addr,
                occupation=occ,
                arrival_date=ad,
                arrival_time=at,
                ward_id=ward_id,
                doctor_id=doctor_id,
                caregiver_exists=caregiver_exists,
                discharge_datetime=discharge_dt or None
            )
            db.session.add(p)
            imported += 1

        db.session.commit()
        flash(f'Imported {imported} patients', 'success')
        return redirect(url_for('settings_import'))

    # GET
    return render_template('settings_import.html', t=t)




# ----------------- Import template (Sample) ----------------

@app.route('/settings/import/template')
@superadmin_required
def settings_import_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Patients"
    headers = ['Ист номер','Фамилия','Имя','Отчество','Дата рождения','Тел номер',
               'Яшаш жойи','Касби','Келган сана','Келган вакти','Палата','Врач','Каровчи']
    ws.append(headers)
    out_path = os.path.join(app.config['UPLOAD_FOLDER'], 'import_template.xlsx')
    wb.save(out_path)
    return send_file(out_path, as_attachment=True, download_name='import_template.xlsx')


# -------------------- Utilities in templates --------------------

@app.context_processor
def inject_utilities():
    # Provide t(), lang, and role flag globally to templates
    user_id = session.get('user_id')
    is_superadmin = False
    if user_id is not None:
        try:
            u = User.query.get(user_id)
            is_superadmin = bool(u and u.role == 'superadmin')
        except Exception:
            is_superadmin = False
    return dict(t=t, lang=get_lang(), is_superadmin=is_superadmin)

# -------------------- Run --------------------
if __name__ == '__main__':
    # Listen on all network interfaces (LAN), production-safe (no debug)
    app.run(host='0.0.0.0', port=5000, debug=False)

